import os
from urllib import request, response

from utils.StringHelper import StringHelper
from utils.HttpMethodsHelper import HttpMethodHelper
from bs4 import BeautifulSoup
import io
from pdfminer.converter import TextConverter
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfpage import PDFPage
from openpyxl import load_workbook, Workbook


class FileHelper:
    @staticmethod
    def transform_link_content_to_txt(link):
        path_to_file = StringHelper.get_random_path("txt")
        with open(path_to_file, "w", encoding='utf-8') as file:
            if 'html' in HttpMethodHelper.get_link_content_type(link):
                print("Andrew here")
                if 'charset' in HttpMethodHelper.get_link_content_type(link):
                    encoding = HttpMethodHelper.get_link_content_encoding(link)
                    print("Andrew here 2")
                else:
                    encoding = None
                    print("Andrew here 3")
                parser = 'html.parser'
                soup = BeautifulSoup(HttpMethodHelper.get_link_content(link),
                                     parser,
                                     from_encoding=encoding)
                file.write(soup.getText())
            elif 'pdf' in HttpMethodHelper.get_link_content_type(link):
                print("Andrew here 345")
                path_to_pdf = StringHelper.get_random_path("pdf")
                with open(path_to_pdf, "wb") as f:
                    f.write(HttpMethodHelper.get_link_content(link))
                    resource_manager = PDFResourceManager()
                    fake_file_handle = io.StringIO()
                    converter = TextConverter(resource_manager,
                                              fake_file_handle)
                    page_interpreter = PDFPageInterpreter(resource_manager,
                                                          converter)

                    with open(path_to_pdf, 'rb') as fh:
                        for page in PDFPage.get_pages(fh,
                                                      caching=True,
                                                      check_extractable=False):
                            page_interpreter.process_page(page)

                        text = fake_file_handle.getvalue()

                    # close open handles
                    converter.close()
                    fake_file_handle.close()

                    if text:
                        file.write(text)
                os.remove(path_to_pdf)
        print("Andrew here finish")
        return path_to_file

    def get_products_from_excel(excel_file):
        wb = load_workbook(excel_file)
        worksheet = wb["Sheet1"]
        excel_data = list()
        col_names = {}
        current = 0
        for col in worksheet.iter_cols(1, worksheet.max_column):
            col_names[col[0].value] = current
            current += 1
        for row_cells in worksheet.iter_rows(min_row=2):
            excel_data.append(row_cells[col_names['Product']].value)
        return excel_data

    def export_products_to_xlsx(devices):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Devices'

        columns = [
            'Название',
            'Среднее время восстановления, ч',
            'Средняя наработка на отказ, ч',
            'Средняя интенсивность отказов, 1 / ч',
            'Интенсивность отказов в режиме хранения',
            'Средний срок сохраняемости, ч',
            'Минимальный ресурс, ч',
            'Гамма-процентный ресурс, ч',
            'Средний ресурс, ч',
            'Средний срок службы, г',
            'Интенсивность восстановления',
            'Надежность системы',
            'Ссылка'
        ]

        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
        
        for device in devices:
            row_num += 1

            row = [
                device.name,
                device.mttr,
                device.mtbf,
                device.failure_rate,
                device.failure_rate_in_storage_mode,
                device.storage_time,
                device.minimal_resource,
                device.gamma_percentage_resource,
                device.average_resource,
                device.average_lifetime,
                device.recovery_intensity,
                device.system_reliability,
                ','.join(device.link),
            ]

            for col_num,cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num,column=col_num)
                cell.value = cell_value
        
        return workbook
